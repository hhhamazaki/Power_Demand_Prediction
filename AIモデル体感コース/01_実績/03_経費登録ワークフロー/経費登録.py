#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
経費登録自動化スクリプト
UiPath Main.xamlをPythonに変換した単一ファイル実装
"""

import os
import time
import logging
import requests
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import re
import tkinter as tk
from tkinter import messagebox
from datetime import datetime

# ====================
# 設定定数
# ====================
DOWNLOAD_URL = "https://www.expense-demo.com/data.xlsx"
EXCEL_FILE_PATH = "./data.xlsx"
SHEET_NAME = "Sheet1"
TARGET_URL = "https://www.expense-demo.com/"
WAIT_TIMEOUT = 10
LOG_FILE_PATH = "./expense_automation.log"
MAX_RETRY_COUNT = 3
RETRY_INTERVAL = 2

# Excelのカラム名
COLUMN_TITLE = "タイトル"
COLUMN_CATEGORY = "種別"
COLUMN_AMOUNT = "金額"
COLUMN_REMARKS = "備考"
COLUMN_NUMBER = "番号"

# ====================
# ログ設定
# ====================
def setup_logging():
    """ログ設定を初期化"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(LOG_FILE_PATH, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

# ====================
# ファイル操作
# ====================
def check_and_download_file():
    """
    data.xlsxファイルの存在確認と必要に応じたダウンロード
    Returns:
        bool: 処理成功フラグ
    """
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            logging.info("ファイルが存在します。")
            show_message("ファイルが存在します。")
            return True
        else:
            logging.info("ファイルが存在しないため、URLからダウンロードします。")
            response = requests.get(DOWNLOAD_URL, timeout=30)
            response.raise_for_status()
            
            with open(EXCEL_FILE_PATH, 'wb') as file:
                file.write(response.content)
            
            logging.info("ファイルをURLからダウンロードしました。")
            show_message("ファイルを URL からダウンロードしました。")
            return True
            
    except Exception as e:
        logging.error(f"ファイルダウンロードエラー: {e}")
        show_message(f"ファイルダウンロードに失敗しました: {e}")
        return False

def read_excel_data():
    """
    Excelファイルからデータを読み取り
    Returns:
        pd.DataFrame: 読み取ったデータ、失敗時はNone
    """
    try:
        df = pd.read_excel(EXCEL_FILE_PATH, sheet_name=SHEET_NAME)
        logging.info(f"Excelファイル読み取り成功。行数: {len(df)}")
        return df
    except Exception as e:
        logging.error(f"Excel読み取りエラー: {e}")
        return None

def update_excel_with_code(row_index, code):
    """
    Excelファイルの指定行の番号列にコードを書き込み
    Args:
        row_index (int): 行インデックス（0ベース）
        code (str): 書き込むコード
    Returns:
        bool: 書き込み成功フラグ
    """
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        worksheet = workbook[SHEET_NAME]
        
        # ヘッダー行を考慮して+2（1ベース + ヘッダー行）
        target_row = row_index + 2
        
        # 番号列を探す（E列と仮定するが、動的に探す）
        header_row = 1
        number_col = None
        for col in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=header_row, column=col).value == COLUMN_NUMBER:
                number_col = col
                break
        
        if number_col:
            worksheet.cell(row=target_row, column=number_col, value=code)
            workbook.save(EXCEL_FILE_PATH)
            logging.info(f"行{row_index + 1}の番号列にコード'{code}'を書き込みました。")
            return True
        else:
            logging.error("番号列が見つかりません。")
            return False
            
    except Exception as e:
        logging.error(f"Excel書き込みエラー: {e}")
        return False

# ====================
# WebDriver関連
# ====================
def initialize_webdriver():
    """
    WebDriverを初期化（複数の方法を試行）
    Returns:
        webdriver: WebDriverインスタンス、失敗時はNone
    """
    driver = None
    
    # 方法1: Selenium 4の内蔵ドライバマネージャ（推奨）
    try:
        options = webdriver.EdgeOptions()
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        driver = webdriver.Edge(options=options)
        logging.info("WebDriver初期化成功（Selenium内蔵マネージャ）")
        
    except Exception as e1:
        logging.warning(f"Selenium内蔵マネージャでの初期化失敗: {e1}")
        
        # 方法2: webdriver-managerを使用した自動ダウンロード
        try:
            from webdriver_manager.microsoft import EdgeChromiumDriverManager
            service = Service(EdgeChromiumDriverManager().install())
            options = webdriver.EdgeOptions()
            options.add_argument("--disable-gpu")
            options.add_argument("--no-sandbox")
            driver = webdriver.Edge(service=service, options=options)
            logging.info("WebDriver初期化成功（webdriver-manager）")
            
        except Exception as e2:
            logging.warning(f"webdriver-manager失敗: {e2}")
            
            # 方法3: ローカルドライバでのフォールバック
            try:
                local_driver_path = "./msedgedriver.exe"
                if os.path.exists(local_driver_path):
                    service = Service(local_driver_path)
                    options = webdriver.EdgeOptions()
                    options.add_argument("--disable-gpu")
                    options.add_argument("--no-sandbox")
                    driver = webdriver.Edge(service=service, options=options)
                    logging.info("WebDriver初期化成功（ローカルドライバ）")
                else:
                    logging.error("ローカルドライバが見つかりません。")
                    
            except Exception as e3:
                logging.error(f"ローカルドライバでの初期化失敗: {e3}")
    
    if driver:
        driver.maximize_window()
        driver.implicitly_wait(5)
    
    return driver

def safe_find_element(driver, selectors, timeout=WAIT_TIMEOUT):
    """
    複数のセレクタを試して要素を安全に検索
    Args:
        driver: WebDriverインスタンス
        selectors: セレクタのリスト [(By.方法, "値"), ...]
        timeout: タイムアウト時間
    Returns:
        WebElement: 見つかった要素、失敗時はNone
    """
    wait = WebDriverWait(driver, timeout)
    
    for by_method, selector_value in selectors:
        try:
            element = wait.until(EC.element_to_be_clickable((by_method, selector_value)))
            return element
        except TimeoutException:
            continue
    
    return None

def safe_click(driver, selectors, description="要素"):
    """
    安全なクリック処理（リトライ付き）
    Args:
        driver: WebDriverインスタンス
        selectors: セレクタのリスト
        description: 要素の説明
    Returns:
        bool: クリック成功フラグ
    """
    for attempt in range(MAX_RETRY_COUNT):
        try:
            element = safe_find_element(driver, selectors)
            if element:
                element.click()
                logging.info(f"{description}をクリックしました。")
                time.sleep(1)  # クリック後の安定化待機
                return True
        except Exception as e:
            logging.warning(f"{description}のクリック試行{attempt + 1}回目失敗: {e}")
            if attempt < MAX_RETRY_COUNT - 1:
                time.sleep(RETRY_INTERVAL)
    
    logging.error(f"{description}のクリックに失敗しました。")
    return False

def safe_input_text(driver, selectors, text, description="入力フィールド"):
    """
    安全なテキスト入力処理
    Args:
        driver: WebDriverインスタンス
        selectors: セレクタのリスト
        text: 入力するテキスト
        description: フィールドの説明
    Returns:
        bool: 入力成功フラグ
    """
    for attempt in range(MAX_RETRY_COUNT):
        try:
            element = safe_find_element(driver, selectors)
            if element:
                element.clear()
                element.send_keys(str(text))
                logging.info(f"{description}に'{text}'を入力しました。")
                return True
        except Exception as e:
            logging.warning(f"{description}の入力試行{attempt + 1}回目失敗: {e}")
            if attempt < MAX_RETRY_COUNT - 1:
                time.sleep(RETRY_INTERVAL)
    
    logging.error(f"{description}への入力に失敗しました。")
    return False

def safe_select_option(driver, selectors, option_text, description="選択フィールド"):
    """
    安全なセレクトボックス選択処理
    Args:
        driver: WebDriverインスタンス
        selectors: セレクタのリスト
        option_text: 選択するオプションのテキスト
        description: フィールドの説明
    Returns:
        bool: 選択成功フラグ
    """
    for attempt in range(MAX_RETRY_COUNT):
        try:
            element = safe_find_element(driver, selectors)
            if element:
                select = Select(element)
                select.select_by_visible_text(str(option_text))
                logging.info(f"{description}で'{option_text}'を選択しました。")
                return True
        except Exception as e:
            logging.warning(f"{description}の選択試行{attempt + 1}回目失敗: {e}")
            if attempt < MAX_RETRY_COUNT - 1:
                time.sleep(RETRY_INTERVAL)
    
    logging.error(f"{description}での選択に失敗しました。")
    return False

# ====================
# 業務処理
# ====================
def navigate_to_expense_site(driver):
    """
    経費登録サイトにアクセス
    Args:
        driver: WebDriverインスタンス
    Returns:
        bool: アクセス成功フラグ
    """
    try:
        driver.get(TARGET_URL)
        logging.info(f"サイトにアクセスしました: {TARGET_URL}")
        return True
    except Exception as e:
        logging.error(f"サイトアクセスエラー: {e}")
        return False

def click_expense_register_menu(driver):
    """
    「経費を登録する」メニューをクリック
    Args:
        driver: WebDriverインスタンス
    Returns:
        bool: クリック成功フラグ
    """
    selectors = [
        (By.LINK_TEXT, "経費を登録する"),
        (By.PARTIAL_LINK_TEXT, "経費を登録"),
        (By.CSS_SELECTOR, "a[href*='expense']"),
        (By.XPATH, "//a[contains(text(), '経費を登録する')]")
    ]
    
    return safe_click(driver, selectors, "経費を登録するメニュー")

def click_new_detail_button(driver):
    """
    「明細を登録する」ボタンをクリック
    Args:
        driver: WebDriverインスタンス
    Returns:
        bool: クリック成功フラグ
    """
    selectors = [
        (By.ID, "newTx"),
        (By.LINK_TEXT, "明細を登録する"),
        (By.PARTIAL_LINK_TEXT, "明細を登録"),
        (By.XPATH, "//a[contains(text(), '明細を登録する')]")
    ]
    
    return safe_click(driver, selectors, "明細を登録するボタン")

def input_form_data(driver, title, category, amount, remarks):
    """
    フォームにデータを入力
    Args:
        driver: WebDriverインスタンス
        title: タイトル
        category: 種別
        amount: 金額
        remarks: 備考
    Returns:
        bool: 入力成功フラグ
    """
    success_count = 0
    
    # タイトル入力
    title_selectors = [
        (By.ID, "ledger_ledger_name"),
        (By.NAME, "ledger[ledger_name]"),
        (By.XPATH, "//input[@placeholder='タイトル']")
    ]
    if safe_input_text(driver, title_selectors, title, "タイトル"):
        success_count += 1
    
    # 種別選択
    category_selectors = [
        (By.TAG_NAME, "select"),
        (By.NAME, "ledger[category]"),
        (By.XPATH, "//select[contains(@name, 'category')]")
    ]
    if safe_select_option(driver, category_selectors, category, "種別"):
        success_count += 1
    
    # 金額入力
    amount_selectors = [
        (By.ID, "ledger_cost"),
        (By.NAME, "ledger[cost]"),
        (By.XPATH, "//input[@placeholder='金額']")
    ]
    if safe_input_text(driver, amount_selectors, amount, "金額"):
        success_count += 1
    
    # 備考入力
    remarks_selectors = [
        (By.ID, "ledger_remarks"),
        (By.NAME, "ledger[remarks]"),
        (By.XPATH, "//input[@placeholder='備考']")
    ]
    if safe_input_text(driver, remarks_selectors, remarks, "備考"):
        success_count += 1
    
    return success_count >= 3  # 最低3項目成功なら OK

def click_register_button(driver):
    """
    「登録する」ボタンをクリック
    Args:
        driver: WebDriverインスタンス
    Returns:
        bool: クリック成功フラグ
    """
    selectors = [
        (By.CSS_SELECTOR, "input[type='submit']"),
        (By.XPATH, "//input[@value='登録する']"),
        (By.XPATH, "//button[contains(text(), '登録する')]")
    ]
    
    return safe_click(driver, selectors, "登録するボタン")

def get_registration_code(driver):
    """
    登録後のコードを取得
    Args:
        driver: WebDriverインスタンス
    Returns:
        str: 取得したコード、失敗時は自動生成コード
    """
    # 複数の方法でコードを取得を試行
    selectors = [
        (By.XPATH, "//dd/strong"),
        (By.XPATH, "//dd[1]"),
        (By.XPATH, "//*[contains(text(), 'コード')]/following-sibling::*"),
        (By.CSS_SELECTOR, "dd"),
        (By.TAG_NAME, "strong")
    ]
    
    for by_method, selector_value in selectors:
        try:
            element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((by_method, selector_value))
            )
            text = element.text.strip()
            # 数字のみ抽出
            code_match = re.search(r'\d{3,}', text)
            if code_match:
                code = code_match.group()
                logging.info(f"登録コードを取得しました: {code}")
                return code
        except:
            continue
    
    # ページソース全体から数字を検索
    try:
        page_source = driver.page_source
        code_matches = re.findall(r'\b\d{6}\b', page_source)
        if code_matches:
            code = code_matches[0]
            logging.info(f"ページソースから登録コードを取得しました: {code}")
            return code
    except:
        pass
    
    # 自動生成フォールバック
    auto_code = f"AUTO_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    logging.warning(f"登録コードの取得に失敗したため、自動生成コードを使用します: {auto_code}")
    return auto_code

def click_back_button(driver):
    """
    「戻る」ボタンをクリック
    Args:
        driver: WebDriverインスタンス
    Returns:
        bool: クリック成功フラグ
    """
    selectors = [
        (By.LINK_TEXT, "戻る"),
        (By.PARTIAL_LINK_TEXT, "戻る"),
        (By.XPATH, "//a[contains(text(), '戻る')]"),
        (By.CSS_SELECTOR, "a[href*='back']")
    ]
    
    return safe_click(driver, selectors, "戻るボタン")

def process_single_row(driver, row_index, row_data):
    """
    1行分の明細登録処理
    Args:
        driver: WebDriverインスタンス
        row_index: 行インデックス
        row_data: 行データ（pandas Series）
    Returns:
        bool: 処理成功フラグ
    """
    try:
        logging.info(f"行{row_index + 1}の処理を開始します。")
        
        # 明細登録ボタンをクリック
        if not click_new_detail_button(driver):
            return False
        
        time.sleep(2)  # ページ遷移待機
        
        # フォームデータ入力
        title = row_data.get(COLUMN_TITLE, "")
        category = row_data.get(COLUMN_CATEGORY, "")
        amount = row_data.get(COLUMN_AMOUNT, "")
        remarks = row_data.get(COLUMN_REMARKS, "")
        
        if not input_form_data(driver, title, category, amount, remarks):
            logging.error(f"行{row_index + 1}: フォーム入力に失敗しました。")
            return False
        
        # 登録ボタンをクリック
        if not click_register_button(driver):
            return False
        
        time.sleep(3)  # 登録処理待機
        
        # 登録コードを取得
        code = get_registration_code(driver)
        
        # Excelに書き戻し
        if not update_excel_with_code(row_index, code):
            logging.error(f"行{row_index + 1}: Excel書き戻しに失敗しました。")
        
        # 戻るボタンをクリック
        if not click_back_button(driver):
            logging.warning(f"行{row_index + 1}: 戻るボタンのクリックに失敗しました。")
        
        time.sleep(2)  # 画面遷移待機
        
        logging.info(f"行{row_index + 1}の処理が完了しました。コード: {code}")
        return True
        
    except Exception as e:
        logging.error(f"行{row_index + 1}の処理中にエラーが発生しました: {e}")
        return False

# ====================
# UI・通知関連
# ====================
def show_message(message):
    """
    メッセージを表示（GUI/CUI対応）
    Args:
        message: 表示するメッセージ
    """
    try:
        # GUIでの表示を試行
        root = tk.Tk()
        root.withdraw()  # メインウィンドウを非表示
        messagebox.showinfo("経費登録自動化", message)
        root.destroy()
    except:
        # CUIでの表示
        print(f"メッセージ: {message}")

def show_completion_message(total_rows, success_count, elapsed_time):
    """
    処理完了メッセージを表示
    Args:
        total_rows: 総行数
        success_count: 成功件数
        elapsed_time: 所要時間（秒）
    """
    message = f"""処理完了

総行数: {total_rows}
成功件数: {success_count}
失敗件数: {total_rows - success_count}
所要時間: {elapsed_time:.1f}秒

OKボタンを押すとブラウザを閉じます。"""
    
    try:
        # GUIでの表示
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("処理完了", message)
        root.destroy()
    except:
        # CUIでの表示
        print("=" * 50)
        print("処理完了")
        print("=" * 50)
        print(f"総行数: {total_rows}")
        print(f"成功件数: {success_count}")
        print(f"失敗件数: {total_rows - success_count}")
        print(f"所要時間: {elapsed_time:.1f}秒")
        print("=" * 50)
        input("Enterキーを押すとブラウザを閉じます...")

# ====================
# メイン処理
# ====================
def main():
    """メイン処理関数"""
    start_time = time.time()
    
    # ログ設定
    setup_logging()
    logging.info("経費登録自動化処理を開始します。")
    
    driver = None
    success_count = 0
    total_rows = 0
    
    try:
        # ステップ1: ファイル確認・ダウンロード
        if not check_and_download_file():
            return
        
        # ステップ2: Excelデータ読み取り
        df = read_excel_data()
        if df is None:
            show_message("Excelファイルの読み取りに失敗しました。")
            return
        
        total_rows = len(df)
        logging.info(f"処理対象行数: {total_rows}")
        
        # ステップ3: WebDriver初期化
        driver = initialize_webdriver()
        if driver is None:
            show_message("WebDriverの初期化に失敗しました。")
            return
        
        # ステップ4: サイトアクセス
        if not navigate_to_expense_site(driver):
            show_message("経費登録サイトへのアクセスに失敗しました。")
            return
        
        # ステップ5: 経費登録メニューをクリック
        if not click_expense_register_menu(driver):
            show_message("経費登録メニューのクリックに失敗しました。")
            return
        
        time.sleep(3)  # ページ遷移待機
        
        # ステップ6: 各行を処理
        for index, row in df.iterrows():
            if process_single_row(driver, index, row):
                success_count += 1
            
            # 行間の待機
            time.sleep(1)
        
        # ステップ7: 完了処理
        elapsed_time = time.time() - start_time
        logging.info(f"全ての処理が完了しました。成功: {success_count}/{total_rows}")
        
        show_completion_message(total_rows, success_count, elapsed_time)
        
    except Exception as e:
        logging.error(f"メイン処理でエラーが発生しました: {e}")
        show_message(f"処理中にエラーが発生しました: {e}")
    
    finally:
        # ブラウザ終了
        if driver:
            try:
                driver.quit()
                logging.info("ブラウザを終了しました。")
            except:
                pass
        
        logging.info("経費登録自動化処理を終了します。")

if __name__ == "__main__":
    main()
