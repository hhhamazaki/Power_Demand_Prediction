import os
import requests
import pandas as pd
import openpyxl
import logging
import time
from datetime import datetime

# WebDriverManagerのログを完全に抑制
import logging
import warnings
warnings.filterwarnings("ignore")
logging.getLogger('WDM').setLevel(logging.CRITICAL)
logging.getLogger('webdriver_manager').setLevel(logging.CRITICAL)
logging.getLogger('selenium').setLevel(logging.ERROR)
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except:
    pass

# tkinter（GUIメッセージボックス）は無い環境もあるため安全にロード
try:
    import tkinter.messagebox as msgbox
except Exception:
    class _DummyMsgBox:
        @staticmethod
        def showinfo(title, message): print(f"[INFO] {title}\n{message}")
        @staticmethod
        def showwarning(title, message): print(f"[WARN] {title}\n{message}")
        @staticmethod
        def showerror(title, message): print(f"[ERROR] {title}\n{message}")
    msgbox = _DummyMsgBox()

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

# コーディング規約 12.2準拠: ログヘッダ定数
LOG_HEADER = "[経費登録_gpt-5.py]"

# ==============================
# 設定管理（コーディング規約 3.4: 設定ファイル活用）
# ==============================

from dataclasses import dataclass, asdict
from typing import Dict, Any, Optional
import json

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

@dataclass
class WebDriverConfig:
    """WebDriver設定"""
    timeout: int = 10
    retry_count: int = 3
    max_window: bool = True

@dataclass
class ApplicationConfig:
    """アプリケーション設定"""
    download_url: str = "https://www.expense-demo.com/data.xlsx"
    target_url: str = "https://www.expense-demo.com/"
    excel_file: str = "data.xlsx"
    sheet_name: str = "Sheet1"

class ConfigurationManager:
    """設定管理サービス（gpt-5特性: 構造化設定活用）"""
    
    def __init__(self, config_file: str = "config.json"):
        self.config_file = os.path.join(BASE_DIR, config_file)
        self._config_cache: Optional[Dict[str, Any]] = None
        self._allow_write = os.getenv('EXPENSE_ALLOW_CONFIG_FILE') == '1'
    
    def load_config(self) -> Dict[str, Any]:
        """設定ファイル読み込み（コーディング規約 3.4準拠）"""
        if self._config_cache:
            return self._config_cache
        
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    self._config_cache = json.load(f)
                    logging.info(f"{LOG_HEADER} 設定ファイル読み込み完了")
            else:
                self._config_cache = self._get_default_config()
                if self._allow_write:
                    self._save_default_config()
                    logging.info(f"{LOG_HEADER} デフォルト設定を作成")
            
            return self._config_cache
            
        except Exception as e:
            logging.warning(f"{LOG_HEADER} 設定ファイル読み込み失敗、デフォルト使用: {e}")
            return self._get_default_config()
    
    def _get_default_config(self) -> Dict[str, Any]:
        """デフォルト設定取得"""
        return {
            "webdriver": asdict(WebDriverConfig()),
            "application": asdict(ApplicationConfig())
        }
    
    def _save_default_config(self) -> None:
        """デフォルト設定保存（許可時のみ）"""
        if not self._allow_write:
            return
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self._config_cache, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.warning(f"{LOG_HEADER} 設定ファイル保存失敗: {e}")

# 設定管理インスタンス
config_manager = ConfigurationManager()
full_config = config_manager.load_config()
app_config = ApplicationConfig(**full_config.get('application', {}))
webdriver_config = WebDriverConfig(**full_config.get('webdriver', {}))

# 設定の辞書形式での提供（既存コードとの互換性のため）
CONFIG = {
    'EXCEL_FILE': app_config.excel_file,
    'SHEET_NAME': app_config.sheet_name,
    'DOWNLOAD_URL': app_config.download_url,
    'TARGET_URL': app_config.target_url,
    'TIMEOUT': webdriver_config.timeout,
    'RETRY_COUNT': webdriver_config.retry_count,
    'MAX_WINDOW': webdriver_config.max_window
}

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(BASE_DIR, 'expense_automation.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# ==============================
# ユーティリティ: 時間表示・メッセージ
# ==============================

class TimeDisplayUtility:
    """時間表示ユーティリティ（gpt-5特性活用）"""
    
    @staticmethod
    def format_duration(start_time: datetime, end_time: datetime) -> str:
        """処理時間のフォーマット"""
        duration = end_time - start_time
        total_seconds = int(duration.total_seconds())
        
        if total_seconds < 60:
            return f"{total_seconds}秒"
        elif total_seconds < 3600:
            minutes = total_seconds // 60
            seconds = total_seconds % 60
            return f"{minutes}分{seconds}秒"
        else:
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return f"{hours}時間{minutes}分{seconds}秒"

def format_time_duration(start_time, end_time):
    """レガシー関数（互換性維持）"""
    return TimeDisplayUtility.format_duration(start_time, end_time)
    duration = end_time - start_time
    total_seconds = int(duration.total_seconds())
    if total_seconds < 60:
        return f"{total_seconds}秒"
    elif total_seconds < 3600:
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes}分{seconds}秒"
    else:
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours}時間{minutes}分{seconds}秒"

def show_processing_summary(start_time, end_time, success_count, total_count, errors=None):
    """処理サマリー"""
    duration_str = format_time_duration(start_time, end_time)
    title = "経費登録処理完了"
    message = (
        f"処理が完了しました。\n\n"
        f"開始時刻: {start_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"終了時刻: {end_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"処理時間: {duration_str}\n\n"
        f"処理結果: {success_count}/{total_count} 件成功"
    )
    if errors:
        trimmed = errors if len(errors) <= 400 else errors[:400] + "..."
        message += f"\n\nエラー詳細:\n{trimmed}"

    try:
        if success_count == total_count:
            msgbox.showinfo(title, message)
        elif success_count > 0:
            msgbox.showwarning(title, message)
        else:
            msgbox.showerror(title, message)
    except Exception:
        print(message)

# ==============================
# ファイル: 存在確認・ダウンロード
# ==============================

def check_and_download_file():
    """ファイル存在確認とダウンロード (UiPath: FileExists → DownloadFileFromUrl)"""
    file_path = CONFIG['EXCEL_FILE']
    try:
        if os.path.exists(file_path):
            logging.info("ファイルが存在します。")
            print("ファイルが存在します。")
            return True

        logging.info("ファイルが存在しないため、ダウンロードを開始します。")
        print("ファイルが存在しないため、ダウンロードを開始します。")
        response = requests.get(CONFIG['DOWNLOAD_URL'], timeout=30)
        response.raise_for_status()

        with open(file_path, 'wb') as f:
            f.write(response.content)

        logging.info("ファイルのダウンロードが完了しました。")
        print("ファイルのダウンロードが完了しました。")
        return True
    except Exception as e:
        error_msg = f"ファイル処理でエラーが発生しました: {e}"
        logging.error(error_msg)
        print(f"エラー: {error_msg}")
        return False

# ==============================
# WebDriver 初期化（Selenium標準機能のみ使用）
# ==============================

def create_edge_driver(options: webdriver.EdgeOptions) -> webdriver.Edge:
    """Selenium Manager（標準機能）のみを使用してEdgeドライバーを取得"""
    try:
        driver = webdriver.Edge(options=options)
        logging.info("Selenium Managerを使用してEdgeドライバーを取得しました。")
        return driver
    except Exception as e:
        logging.error(f"Edgeドライバーの初期化に失敗しました: {e}")
        raise

# ==============================
# 安全なWeb操作関数
# ==============================

def safe_click(driver, wait, locator, element_name="要素", timeout=None):
    if timeout is None:
        timeout = CONFIG['TIMEOUT']
    try:
        element = wait.until(EC.element_to_be_clickable(locator))
        element.click()
        logging.info(f"{element_name} をクリックしました")
        return True
    except TimeoutException:
        logging.error(f"{element_name} が見つからないか、クリックできませんでした (タイムアウト)")
        return False
    except Exception as e:
        logging.error(f"{element_name} のクリックでエラー: {e}")
        return False

def safe_input(driver, wait, locator, text, element_name="入力フィールド", timeout=None):
    if timeout is None:
        timeout = CONFIG['TIMEOUT']
    try:
        element = wait.until(EC.presence_of_element_located(locator))
        element.clear()
        element.send_keys("" if text is None else str(text))
        logging.info(f"{element_name} に '{text}' を入力しました")
        return True
    except TimeoutException:
        logging.error(f"{element_name} が見つかりませんでした (タイムアウト)")
        return False
    except Exception as e:
        logging.error(f"{element_name} の入力でエラー: {e}")
        return False

def safe_select(driver, wait, locator, value, element_name="選択フィールド", timeout=None):
    if timeout is None:
        timeout = CONFIG['TIMEOUT']
    try:
        select_element = wait.until(EC.presence_of_element_located(locator))
        Select(select_element).select_by_visible_text("" if value is None else str(value))
        logging.info(f"{element_name} で '{value}' を選択しました")
        return True
    except TimeoutException:
        logging.error(f"{element_name} が見つかりませんでした (タイムアウト)")
        return False
    except Exception as e:
        logging.error(f"{element_name} の選択でエラー: {e}")
        return False

def get_text_safe(driver, wait, locator, element_name="テキスト要素", timeout=None):
    if timeout is None:
        timeout = CONFIG['TIMEOUT']
    try:
        element = wait.until(EC.presence_of_element_located(locator))
        text = element.text.strip()
        logging.info(f"{element_name} からテキスト '{text}' を取得しました")
        return text
    except TimeoutException:
        logging.error(f"{element_name} が見つかりませんでした (タイムアウト)")
        return None
    except Exception as e:
        logging.error(f"{element_name} のテキスト取得でエラー: {e}")
        return None

def get_text_safe_quick(driver, wait, locator, element_name="テキスト要素", timeout=5):
    """短時間でテキスト取得を試行"""
    try:
        element = wait.until(EC.presence_of_element_located(locator))
        text = element.text.strip()
        if text:
            logging.info(f"{element_name} からテキスト '{text}' を取得しました")
            return text
        return None
    except TimeoutException:
        return None
    except Exception as e:
        logging.error(f"{element_name} のテキスト取得でエラー: {e}")
        return None

# ==============================
# 経費登録1件分の処理
# ==============================

def register_expense_item(driver, wait, title, category, amount, remarks):
    """1行分の明細登録"""
    try:
        logging.info(f"経費項目登録開始: タイトル='{title}', 種別='{category}', 金額='{amount}', 備考='{remarks}'")

        # 「明細を登録する」クリック（ID優先 → テキスト）
        if not safe_click(driver, wait, (By.ID, "newTx"), "明細を登録するボタン"):
            if not safe_click(driver, wait, (By.LINK_TEXT, "明細を登録する"), "明細を登録するボタン"):
                raise Exception("明細を登録するボタンが見つかりません")

        # タイトル
        if not safe_input(driver, wait, (By.ID, "ledger_ledger_name"), title, "タイトルフィールド"):
            raise Exception("タイトルの入力に失敗しました")

        # 種別 (より効率的なロケータ順序)
        select_locators = [
            (By.CSS_SELECTOR, "select"),
            (By.XPATH, "//select[1]"),
            (By.XPATH, "//select[@id='ledger_type' or @name='ledger[type]' or @name='type' or @name='category']"),
            (By.XPATH, "//select[@aria-label='種別' or @title='種別' or @data-name='種別']"),
        ]
        selected = False
        for loc in select_locators:
            if safe_select(driver, wait, loc, category, "種別フィールド", timeout=10):
                selected = True
                break
        if not selected:
            logging.warning("種別選択に失敗しましたが、処理を継続します")

        # 金額
        if not safe_input(driver, wait, (By.ID, "ledger_cost"), amount, "金額フィールド"):
            raise Exception("金額の入力に失敗しました")

        # 備考（任意）
        if not safe_input(driver, wait, (By.ID, "ledger_remarks"), remarks, "備考フィールド"):
            logging.warning("備考の入力に失敗しましたが、処理を継続します")

        # 「登録する」クリック
        submit_locators = [
            (By.XPATH, "//input[@type='submit' and (@value='登録する' or @value='Submit')]"),
            (By.CSS_SELECTOR, "input[type='submit']"),
            (By.XPATH, "//button[normalize-space()='登録する']")
        ]
        clicked = any(safe_click(driver, wait, loc, "登録ボタン") for loc in submit_locators)
        if not clicked:
            raise Exception("登録ボタンのクリックに失敗しました")

        # 完了画面でコード取得 (短時間で複数ロケータを試行)
        code_locators = [
            (By.XPATH, "//dd[1]"),
            (By.XPATH, "//strong[text()='コード:']/following-sibling::*[1]"),
            (By.XPATH, "//dt[contains(text(), 'コード')]/following-sibling::dd[1]"),
        ]
        registration_code = None
        for loc in code_locators:
            registration_code = get_text_safe_quick(driver, wait, loc, "登録コード", timeout=5)
            if registration_code:
                break
        if not registration_code:
            logging.warning("登録コードの取得に失敗しました。タイムスタンプを代替値として使用します")
            registration_code = f"AUTO_{int(time.time())}"

        # 「戻る」クリック
        back_locators = [
            (By.LINK_TEXT, "戻る"),
            (By.XPATH, "//a[normalize-space()='戻る']"),
            (By.XPATH, "//button[normalize-space()='戻る']")
        ]
        any(safe_click(driver, wait, loc, "戻るボタン") for loc in back_locators)

        logging.info(f"経費項目登録完了: 登録コード='{registration_code}'")
        return registration_code

    except Exception as e:
        logging.error(f"経費項目登録エラー: {e}")
        return None

# ==============================
# Excel各行の処理
# ==============================

def process_excel_rows(driver, wait, df, worksheet):
    total_rows = len(df)
    success_count = 0
    error_messages = []

    logging.info(f"処理対象行数: {total_rows}")

    for index, row in df.iterrows():
        excel_row = index + 2  # ヘッダ行を考慮
        try:
            title = str(row['タイトル']) if pd.notna(row['タイトル']) else ""
            category = str(row['種別']) if pd.notna(row['種別']) else ""
            amount = str(row['金額']) if pd.notna(row['金額']) else ""
            remarks = str(row['備考']) if pd.notna(row['備考']) else ""

            logging.info(f"行 {excel_row} の処理開始 ({index + 1}/{total_rows})")

            registration_code = register_expense_item(driver, wait, title, category, amount, remarks)
            if registration_code:
                worksheet.cell(row=excel_row, column=5, value=registration_code)  # E列 = 番号
                success_count += 1
                logging.info(f"行 {excel_row} 完了: 登録コード='{registration_code}'")
            else:
                msg = f"行 {excel_row} の処理失敗"
                logging.error(msg)
                error_messages.append(msg)
                worksheet.cell(row=excel_row, column=5, value="ERROR")

        except Exception as e:
            msg = f"行 {excel_row} の処理でエラー: {e}"
            logging.error(msg)
            error_messages.append(msg)
            worksheet.cell(row=excel_row, column=5, value="ERROR")
            continue

    logging.info(f"処理完了: 成功 {success_count}/{total_rows} 行")
    return success_count, error_messages

# ==============================
# Web処理メイン
# ==============================

def web_processing(df, workbook, worksheet):
    driver = None
    start_time = datetime.now()

    try:
        logging.info("WebDriver を初期化しています...")
        options = webdriver.EdgeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        options.add_experimental_option('useAutomationExtension', False)

        driver = create_edge_driver(options)
        wait = WebDriverWait(driver, CONFIG['TIMEOUT'])

        # サイトアクセス
        logging.info(f"サイトにアクセス: {CONFIG['TARGET_URL']}")
        driver.maximize_window()
        driver.get(CONFIG['TARGET_URL'])

        # 「経費を登録する」クリック (複数ロケータ)
        expense_locators = [
            (By.LINK_TEXT, "経費を登録する"),
            (By.XPATH, "//a[normalize-space()='経費を登録する']"),
            (By.XPATH, "//a[contains(@class,'menu-link') and normalize-space()='経費を登録する']")
        ]
        if not any(safe_click(driver, wait, loc, "経費を登録するボタン") for loc in expense_locators):
            raise Exception("経費を登録するボタンが見つかりません")

        # 各行処理
        success_count, error_messages = process_excel_rows(driver, wait, df, worksheet)

        # サマリー表示
        end_time = datetime.now()
        error_summary = "; ".join(error_messages[:3]) if error_messages else None
        show_processing_summary(start_time, end_time, success_count, len(df), error_summary)
        return success_count

    except Exception as e:
        end_time = datetime.now()
        logging.error(f"Web処理エラー: {e}")
        show_processing_summary(start_time, end_time, 0, len(df), str(e))
        raise
    finally:
        if driver:
            try:
                driver.quit()
                logging.info("WebDriver を終了しました")
            except Exception:
                pass

# ==============================
# Excel処理メイン
# ==============================

def excel_processing():
    try:
        logging.info("Excel処理を開始しています...")
        df = pd.read_excel(CONFIG['EXCEL_FILE'], sheet_name=CONFIG['SHEET_NAME'])
        logging.info(f"Excel データを読み込みました: {len(df)} 行")

        if df.empty:
            raise Exception("Excel ファイルにデータがありません")

        # 必須列確認
        required = ['タイトル', '種別', '金額', '備考']
        miss = [c for c in required if c not in df.columns]
        if miss:
            raise Exception(f"必要な列が見つかりません: {miss}")

        # 書き込み用にopenpyxlで開く
        workbook = openpyxl.load_workbook(CONFIG['EXCEL_FILE'])
        worksheet = workbook[CONFIG['SHEET_NAME']]

        # 「番号」ヘッダが未設定なら補完 (E列)
        if worksheet.cell(row=1, column=5).value != '番号':
            worksheet.cell(row=1, column=5, value='番号')

        # Web操作
        success_count = web_processing(df, workbook, worksheet)

        # 保存
        workbook.save(CONFIG['EXCEL_FILE'])
        workbook.close()

        logging.info(f"Excel処理完了: {success_count} 件成功")
        return True

    except FileNotFoundError:
        msg = f"ファイルが見つかりません: {CONFIG['EXCEL_FILE']}"
        logging.error(msg)
        print(f"エラー: {msg}")
        return False
    except Exception as e:
        msg = f"Excel処理エラー: {e}"
        logging.error(msg)
        print(f"エラー: {msg}")
        return False

# ==============================
# テスト機能
# ==============================

def test_basic_functionality():
    print("=== 基本機能テスト開始 ===")
    # 1. ライブラリインポート
    try:
        print("1. ライブラリインポート...")
        import pandas as _pd
        import openpyxl as _ox
        from selenium import webdriver as _wd
        print("   ✓ 成功")
    except Exception as e:
        print(f"   ✗ 失敗: {e}")
        return False

    # 2. ファイル存在確認＆DL
    try:
        print("2. ファイル準備...")
        if check_and_download_file():
            print("   ✓ 成功")
        else:
            print("   ✗ 失敗")
            return False
    except Exception as e:
        print(f"   ✗ 失敗: {e}")
        return False

    # 3. Excel読込
    try:
        print("3. Excel読込...")
        df = pd.read_excel(CONFIG['EXCEL_FILE'], sheet_name=CONFIG['SHEET_NAME'])
        print(f"   ✓ {len(df)} 行読み込み")
        print(f"   ✓ 列: {list(df.columns)}")
    except Exception as e:
        print(f"   ✗ 失敗: {e}")
        return False

    # 4. WebDriver初期化（ヘッドレスで簡易検証）
    try:
        print("4. WebDriver初期化 (headless)...")
        options = webdriver.EdgeOptions()
        options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        driver = create_edge_driver(options)
        driver.quit()
        print("   ✓ 成功")
    except Exception as e:
        print(f"   ✗ 失敗: {e}")
        print("   → オフライン/社内プロキシ環境では drivers/msedgedriver/msedgedriver.exe を配置してください。")
        return False

    print("=== 基本機能テスト完了 ===")
    return True

# ==============================
# メイン
# ==============================

def cleanup_config_file():
    """実行完了後のconfig.jsonクリーンアップ（許可時のみ）"""
    try:
        if os.getenv('EXPENSE_ALLOW_CONFIG_FILE') == '1':
            config_file_path = os.path.join(BASE_DIR, "config.json")
            if os.path.exists(config_file_path):
                os.remove(config_file_path)
                logging.info(f"{LOG_HEADER} 設定ファイルを削除しました: config.json")
    except Exception as e:
        logging.warning(f"{LOG_HEADER} 設定ファイル削除に失敗: {e}")

def main():
    import sys
    # テストモード
    if len(sys.argv) > 1 and sys.argv[1].lower() == "test":
        return test_basic_functionality()

    overall_start = datetime.now()
    try:
        logging.info("=== 経費登録自動化処理開始 ===")
        print(f"経費登録自動化処理を開始します。開始時刻: {overall_start.strftime('%Y-%m-%d %H:%M:%S')}")

        # ファイル準備
        if not check_and_download_file():
            logging.error("ファイル準備に失敗しました")
            return False

        # Excel・Web処理
        ok = excel_processing()
        overall_end = datetime.now()
        duration_str = format_time_duration(overall_start, overall_end)

        if ok:
            msg = f"処理が正常に完了しました。\n総処理時間: {duration_str}"
            print(msg)
            logging.info("=== 経費登録自動化処理完了 ===")
            logging.info(f"終了時刻: {overall_end.strftime('%Y-%m-%d %H:%M:%S')}")
            logging.info(f"総処理時間: {duration_str}")
            
            # 正常完了後にconfig.jsonを削除
            cleanup_config_file()
            return True
        else:
            err = f"Excel・Web処理に失敗しました。\n処理時間: {duration_str}"
            logging.error(err)
            try:
                msgbox.showerror("処理失敗", err)
            except Exception:
                print(f"エラー: {err}")
            return False

    except Exception as e:
        overall_end = datetime.now()
        duration_str = format_time_duration(overall_start, overall_end)
        err = f"処理中にエラーが発生しました: {e}\n処理時間: {duration_str}"
        logging.error(err)
        try:
            msgbox.showerror("処理エラー", err)
        except Exception:
            print(f"エラー: {err}")
        return False
    
    finally:
        # 実行完了後にconfig.jsonを削除
        cleanup_config_file()

if __name__ == "__main__":
    main()
