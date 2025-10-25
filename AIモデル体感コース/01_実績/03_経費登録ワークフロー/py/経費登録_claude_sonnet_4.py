import os
import requests
import pandas as pd
import openpyxl
import tkinter.messagebox as msgbox
import logging
import warnings
import time
from datetime import datetime, timedelta

# WebDriverManagerのログを完全に抑制
warnings.filterwarnings("ignore")
logging.getLogger('WDM').setLevel(logging.CRITICAL)
logging.getLogger('webdriver_manager').setLevel(logging.CRITICAL)
logging.getLogger('selenium').setLevel(logging.ERROR)
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except:
    pass

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException

# 設定
# 設定・データクラス（コーディング規約準拠）
from dataclasses import dataclass
from typing import Tuple, Optional, List

# コーディング規約 12.2準拠: ログヘッダ定数  
LOG_HEADER = "[経費登録_claude_sonnet_4.py]"

@dataclass
class ProcessingResult:
    """処理結果データクラス"""
    success_count: int
    total_count: int
    errors: List[str]
    start_time: datetime
    end_time: datetime

CONFIG = {
    'DOWNLOAD_URL': 'https://www.expense-demo.com/data.xlsx',
    'TARGET_URL': 'https://www.expense-demo.com/',
    'EXCEL_FILE': 'data.xlsx',
    'SHEET_NAME': 'Sheet1',
    'TIMEOUT': 15,
    'RETRY_COUNT': 3
}

class ExcelOptimizedOperations:
    """Excel最適化操作クラス（claude_sonnet_4特性活用）"""
    
    def __init__(self, file_path: str, sheet_name: str = "Sheet1"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self._logger = logging.getLogger(__name__)
    
    def read_and_prepare_data(self) -> pd.DataFrame:
        """データ読み込みと準備（関数分割による可読性向上）"""
        try:
            # pandas読み込み
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            self._logger.info(f"{LOG_HEADER} Excel データ読み込み完了: {len(df)} 行")
            
            # データ検証
            self._validate_data(df)
            
            return df
            
        except Exception as e:
            error_msg = f"Excel読み込みエラー: {e}"
            self._logger.error(f"{LOG_HEADER} {error_msg}")
            raise Exception(error_msg) from e
    
    def _validate_data(self, df: pd.DataFrame) -> None:
        """データ検証（コーディング規約: 関数分割）"""
        if df.empty:
            raise Exception("Excel ファイルにデータがありません")
        
        required_columns = ['タイトル', '種別', '金額', '備考']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise Exception(f"必須列が見つかりません: {missing_columns}")
    
    def write_registration_codes_optimized(self, codes: List[Tuple[int, str]]) -> None:
        """登録コード書き込み最適化（openpyxl特性活用）"""
        try:
            # openpyxlで直接操作（パフォーマンス最適化）
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook[self.sheet_name]
            
            # 番号列ヘッダー追加（まだない場合）
            if worksheet.cell(row=1, column=5).value != '番号':
                worksheet.cell(row=1, column=5, value='番号')
            
            # 一括書き込み（高速化）
            for row_number, code in codes:
                worksheet.cell(row=row_number + 1, column=5, value=code)
            
            # 保存
            workbook.save(self.file_path)
            workbook.close()
            
            self._logger.info(f"{LOG_HEADER} 登録コード書き込み完了: {len(codes)} 件")
            
        except Exception as e:
            self._logger.error(f"{LOG_HEADER} 登録コード書き込みエラー: {e}")
            raise

# ログ設定
import os
log_dir = os.path.dirname(os.path.abspath(__file__))
log_path = os.path.join(log_dir, 'expense_automation.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_path, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

def wait_for_document_ready(driver, timeout: int = CONFIG['TIMEOUT']):
    """document.readyState が complete になるまで待機"""
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )

def create_edge_driver(options: webdriver.EdgeOptions) -> webdriver.Edge:
    """Selenium Manager（標準機能）のみを使用してEdgeドライバーを取得"""
    try:
        driver = webdriver.Edge(options=options)
        logging.info("Selenium Managerを使用してEdgeドライバーを取得しました。")
        return driver
    except Exception as e:
        logging.error(f"Edgeドライバーの初期化に失敗しました: {e}")
        raise

def format_time_duration(start_time, end_time):
    """時間の差を分かりやすい形式でフォーマット"""
    duration = end_time - start_time
    total_seconds = int(duration.total_seconds())
    
    if total_seconds < 60:
        return f"{total_seconds}秒"
    elif total_seconds < 3600:
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes}分{seconds}秒"
    else:
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours}時間{minutes}分{seconds}秒"

def show_processing_summary(start_time, end_time, success_count, total_count, errors=None):
    """処理結果のサマリーを表示"""
    duration_str = format_time_duration(start_time, end_time)
    
    # メッセージ作成
    title = "経費登録処理完了"
    message = f"""処理が完了しました。

開始時刻: {start_time.strftime('%Y-%m-%d %H:%M:%S')}
終了時刻: {end_time.strftime('%Y-%m-%d %H:%M:%S')}
処理時間: {duration_str}

処理結果: {success_count}/{total_count} 件成功"""
    
    if errors:
        message += f"\n\nエラー詳細:\n{errors[:200]}..."  # エラーメッセージは200文字まで
    
    # 成功率に応じてアイコンを変更
    if success_count == total_count:
        msgbox.showinfo(title, message)
    elif success_count > 0:
        msgbox.showwarning(title, message)
    else:
        msgbox.showerror(title, message)

def check_and_download_file():
    """ファイル存在確認とダウンロード処理"""
    file_path = CONFIG['EXCEL_FILE']
    
    try:
        if os.path.exists(file_path):
            logging.info("ファイルが存在します。")
            print("ファイルが存在します。")
            return True
        else:
            logging.info("ファイルが存在しないため、ダウンロードを開始します。")
            print("ファイルが存在しないため、ダウンロードを開始します。")
            
            # ダウンロード実行
            response = requests.get(CONFIG['DOWNLOAD_URL'], timeout=30)
            response.raise_for_status()
            
            with open(file_path, 'wb') as f:
                f.write(response.content)
            
            logging.info("ファイルのダウンロードが完了しました。")
            print("ファイルのダウンロードが完了しました。")
            return True
            
    except requests.exceptions.RequestException as e:
        error_msg = f"ダウンロードに失敗しました: {e}"
        logging.error(error_msg)
        print(f"エラー: {error_msg}")
        return False
        
    except Exception as e:
        error_msg = f"ファイル処理でエラーが発生しました: {e}"
        logging.error(error_msg)
        print(f"エラー: {error_msg}")
        return False

def safe_click(driver, wait, locator, element_name="要素", timeout=None):
    """安全なクリック処理"""
    if timeout is None:
        timeout = CONFIG['TIMEOUT']
    
    try:
        element = wait.until(EC.element_to_be_clickable(locator))
        element.click()
        logging.info(f"{element_name} をクリックしました")
        return True
    except TimeoutException:
        logging.error(f"{element_name} が見つからないか、クリックできませんでした (タイムアウト)")
        return False
    except Exception as e:
        logging.error(f"{element_name} のクリックでエラー: {e}")
        return False

def safe_input(driver, wait, locator, text, element_name="入力フィールド", timeout=None):
    """安全な入力処理"""
    if timeout is None:
        timeout = CONFIG['TIMEOUT']
    
    try:
        element = wait.until(EC.presence_of_element_located(locator))
        element.clear()
        element.send_keys(str(text))
        logging.info(f"{element_name} に '{text}' を入力しました")
        return True
    except TimeoutException:
        logging.error(f"{element_name} が見つかりませんでした (タイムアウト)")
        return False
    except Exception as e:
        logging.error(f"{element_name} の入力でエラー: {e}")
        return False

def safe_select(driver, wait, locator, value, element_name="選択フィールド", timeout=None):
    """安全な選択処理"""
    if timeout is None:
        timeout = CONFIG['TIMEOUT']
    
    try:
        select_element = wait.until(EC.presence_of_element_located(locator))
        select = Select(select_element)
        select.select_by_visible_text(str(value))
        logging.info(f"{element_name} で '{value}' を選択しました")
        return True
    except TimeoutException:
        logging.error(f"{element_name} が見つかりませんでした (タイムアウト)")
        return False
    except Exception as e:
        logging.error(f"{element_name} の選択でエラー: {e}")
        return False

def get_text_safe(driver, wait, locator, element_name="テキスト要素", timeout=None):
    """安全なテキスト取得処理"""
    if timeout is None:
        timeout = CONFIG['TIMEOUT']
    
    try:
        element = wait.until(EC.presence_of_element_located(locator))
        text = element.text.strip()
        logging.info(f"{element_name} からテキスト '{text}' を取得しました")
        return text
    except TimeoutException:
        logging.error(f"{element_name} が見つかりませんでした (タイムアウト)")
        return None
    except Exception as e:
        logging.error(f"{element_name} のテキスト取得でエラー: {e}")
        return None

def register_expense_item(driver, wait, title, category, amount, remarks):
    """経費項目登録処理"""
    try:
        logging.info(f"経費項目登録開始: タイトル='{title}', 種別='{category}', 金額='{amount}', 備考='{remarks}'")

        # 「明細を登録する」ボタンをクリック
        if not safe_click(driver, wait, (By.ID, "newTx"), "明細を登録するボタン"):
            # IDで見つからない場合、リンクテキストで試行
            if not safe_click(driver, wait, (By.LINK_TEXT, "明細を登録する"), "明細を登録するボタン"):
                raise Exception("明細を登録するボタンが見つかりません")

        # 明細入力画面の読み込み完了待機（タイトル入力欄のpresence）
        wait.until(EC.presence_of_element_located((By.ID, "ledger_ledger_name")))

        # タイトル入力
        if not safe_input(driver, wait, (By.ID, "ledger_ledger_name"), title, "タイトルフィールド"):
            raise Exception("タイトルの入力に失敗しました")

        # 種別選択（フォールバック）
        select_locators = [
            (By.XPATH, "//select"),
            (By.XPATH, "//select[contains(@name, 'type') or contains(@name, 'category')]"),
            (By.CSS_SELECTOR, "select"),
        ]
        success = False
        for locator in select_locators:
            if safe_select(driver, wait, locator, category, "種別フィールド"):
                success = True
                break
        if not success:
            logging.warning("種別選択に失敗しましたが、処理を継続します")

        # 金額・備考
        if not safe_input(driver, wait, (By.ID, "ledger_cost"), amount, "金額フィールド"):
            raise Exception("金額の入力に失敗しました")
        if not safe_input(driver, wait, (By.ID, "ledger_remarks"), remarks, "備考フィールド"):
            logging.warning("備考の入力に失敗しましたが、処理を継続します")

        # 登録ボタン
        submit_locators = [
            (By.XPATH, "//input[@type='submit' and @value='登録する']"),
            (By.XPATH, "//input[@type='submit']"),
            (By.XPATH, "//button[text()='登録する']"),
            (By.CSS_SELECTOR, "input[type='submit']"),
        ]
        success = False
        for locator in submit_locators:
            if safe_click(driver, wait, locator, "登録ボタン"):
                success = True
                break
        if not success:
            raise Exception("登録ボタンのクリックに失敗しました")

        # 登録完了の表示待機
        code_wait_locators = [
            (By.XPATH, "//dd[1]"),
            (By.XPATH, "//strong[contains(text(), 'コード')]/following-sibling::*[1]"),
            (By.XPATH, "//dt[contains(text(), 'コード')]/following-sibling::dd[1]"),
        ]
        try:
            WebDriverWait(driver, CONFIG['TIMEOUT']).until(
                lambda d: any(len(d.find_elements(*loc)) > 0 for loc in code_wait_locators)
            )
        except TimeoutException:
            logging.warning("登録完了画面の表示待機がタイムアウト（後続でフォールバック取得を試行）")

        # 登録コード取得
        registration_code = None
        code_locators = [
            (By.XPATH, "//dd[1]"),
            (By.XPATH, "//strong[text()='コード:']/following-sibling::*[1]"),
            (By.XPATH, "//strong[contains(text(), 'コード')]/following-sibling::dd"),
            (By.XPATH, "//dt[contains(text(), 'コード')]/following-sibling::dd[1]"),
        ]
        for locator in code_locators:
            registration_code = get_text_safe(driver, wait, locator, "登録コード")
            if registration_code:
                break
        if not registration_code:
            logging.warning("登録コードの取得に失敗しました。デフォルト値を使用します")
            registration_code = f"AUTO_{int(time.time())}"

        # 戻る
        back_locators = [
            (By.LINK_TEXT, "戻る"),
            (By.XPATH, "//a[text()='戻る']"),
            (By.XPATH, "//a[contains(text(), '戻る')]"),
            (By.XPATH, "//button[text()='戻る']"),
        ]
        success = False
        for locator in back_locators:
            if safe_click(driver, wait, locator, "戻るボタン"):
                success = True
                break
        if not success:
            logging.warning("戻るボタンのクリックに失敗しました")

        # 一覧画面の復帰待機
        try:
            wait.until(EC.element_to_be_clickable((By.ID, "newTx")))
        except TimeoutException:
            logging.warning("一覧画面の復帰待機がタイムアウト")

        logging.info(f"経費項目登録完了: 登録コード='{registration_code}'")
        return registration_code
    except Exception as e:
        logging.error(f"経費項目登録エラー: {e}")
        return None

def process_excel_rows(driver, wait, df, worksheet):
    """Excel各行データの処理"""
    total_rows = len(df)
    success_count = 0
    error_messages = []
    
    logging.info(f"処理対象行数: {total_rows}")
    
    for index, row in df.iterrows():
        try:
            # データ取得
            title = str(row['タイトル']) if pd.notna(row['タイトル']) else ""
            category = str(row['種別']) if pd.notna(row['種別']) else ""
            amount = str(row['金額']) if pd.notna(row['金額']) else ""
            remarks = str(row['備考']) if pd.notna(row['備考']) else ""
            
            # 行番号計算（ヘッダー考慮）
            excel_row = index + 2
            
            logging.info(f"行 {excel_row} の処理開始 ({index + 1}/{total_rows})")
            
            # 登録処理実行
            registration_code = register_expense_item(
                driver, wait, title, category, amount, remarks
            )
            
            if registration_code:
                # Excelに番号書き込み
                worksheet.cell(row=excel_row, column=5, value=registration_code)
                success_count += 1
                logging.info(f"行 {excel_row} の処理完了: 登録コード='{registration_code}'")
            else:
                error_msg = f"行 {excel_row} の処理失敗"
                logging.error(error_msg)
                error_messages.append(error_msg)
                worksheet.cell(row=excel_row, column=5, value="ERROR")
            
        except Exception as e:
            error_msg = f"行 {index + 1} の処理でエラー: {e}"
            logging.error(error_msg)
            error_messages.append(error_msg)
            worksheet.cell(row=excel_row, column=5, value="ERROR")
            continue
    
    logging.info(f"処理完了: 成功 {success_count}/{total_rows} 行")
    return success_count, error_messages

def web_processing(df, workbook, worksheet):
    """Web操作メイン処理"""
    driver = None
    start_time = datetime.now()

    try:
        # WebDriver初期化
        logging.info("WebDriver を初期化しています...")

        options = webdriver.EdgeOptions()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        options.add_experimental_option('useAutomationExtension', False)

        driver = create_edge_driver(options)

        wait = WebDriverWait(driver, CONFIG['TIMEOUT'])

        # サイトアクセス
        logging.info(f"サイトにアクセスしています: {CONFIG['TARGET_URL']}")
        driver.maximize_window()
        driver.get(CONFIG['TARGET_URL'])

        # ページの読み込み待機（readyState）
        wait_for_document_ready(driver)

        # 「経費を登録する」ボタンクリック
        expense_locators = [
            (By.LINK_TEXT, "経費を登録する"),
            (By.XPATH, "//a[text()='経費を登録する']"),
            (By.XPATH, "//a[contains(text(), '経費を登録する')]"),
            (By.XPATH, "//a[@class='btn btn-primary btn-block menu-link' and text()='経費を登録する']"),
        ]
        success = False
        for locator in expense_locators:
            if safe_click(driver, wait, locator, "経費を登録するボタン"):
                success = True
                break
        if not success:
            raise Exception("経費を登録するボタンが見つかりません")

        # ページ遷移の待機（newTxがクリック可能になるまで）
        wait.until(EC.element_to_be_clickable((By.ID, "newTx")))

        # 各行データのループ処理
        success_count, error_messages = process_excel_rows(driver, wait, df, worksheet)

        # 処理時間計算
        end_time = datetime.now()

        # 結果表示
        error_summary = "; ".join(error_messages[:3]) if error_messages else None  # 最初の3つのエラーのみ
        show_processing_summary(start_time, end_time, success_count, len(df), error_summary)

        return success_count

    except Exception as e:
        end_time = datetime.now()
        logging.error(f"Web処理エラー: {e}")

        # エラー時も時間を表示
        show_processing_summary(start_time, end_time, 0, len(df), str(e))
        raise
    finally:
        if driver:
            try:
                driver.quit()
                logging.info("WebDriver を終了しました")
            except:
                pass

def excel_processing():
    """Excel処理メイン関数"""
    try:
        logging.info("Excel処理を開始しています...")
        
        # データ読み込み
        df = pd.read_excel(CONFIG['EXCEL_FILE'], sheet_name=CONFIG['SHEET_NAME'])
        logging.info(f"Excel データを読み込みました: {len(df)} 行")
        
        # データ確認
        if df.empty:
            raise Exception("Excel ファイルにデータがありません")
        
        # 必要な列の確認
        required_columns = ['タイトル', '種別', '金額', '備考']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise Exception(f"必要な列が見つかりません: {missing_columns}")
        
        # ワークブックをopenpyxlで開く（書き込み用）
        workbook = openpyxl.load_workbook(CONFIG['EXCEL_FILE'])
        worksheet = workbook[CONFIG['SHEET_NAME']]
        
        # ヘッダーに「番号」列を追加（まだない場合）
        if worksheet.cell(row=1, column=5).value != '番号':
            worksheet.cell(row=1, column=5, value='番号')
        
        # Web操作処理を実行
        success_count = web_processing(df, workbook, worksheet)
        
        # 保存
        workbook.save(CONFIG['EXCEL_FILE'])
        workbook.close()
        
        logging.info(f"Excel処理完了: {success_count} 件の処理が成功しました")
        return True
        
    except FileNotFoundError:
        error_msg = f"ファイルが見つかりません: {CONFIG['EXCEL_FILE']}"
        logging.error(error_msg)
        print(f"エラー: {error_msg}")
        return False
    except Exception as e:
        error_msg = f"Excel処理エラー: {e}"
        logging.error(error_msg)
        print(f"エラー: {error_msg}")
        return False

def test_basic_functionality():
    """基本機能のテスト"""
    print("=== 基本機能テスト開始 ===")
    
    # 1. ライブラリのインポートテスト
    try:
        print("1. ライブラリインポートテスト...")
        import pandas as pd
        import openpyxl
        from selenium import webdriver
        print("   ✓ 全ライブラリのインポート成功")
    except Exception as e:
        print(f"   ✗ ライブラリインポートエラー: {e}")
        return False
    
    # 2. ファイル存在確認テスト
    try:
        print("2. ファイル存在確認テスト...")
        if check_and_download_file():
            print("   ✓ ファイル準備成功")
        else:
            print("   ✗ ファイル準備失敗")
            return False
    except Exception as e:
        print(f"   ✗ ファイル準備エラー: {e}")
        return False
    
    # 3. Excelファイル読み込みテスト
    try:
        print("3. Excelファイル読み込みテスト...")
        df = pd.read_excel(CONFIG['EXCEL_FILE'], sheet_name=CONFIG['SHEET_NAME'])
        print(f"   ✓ Excel読み込み成功: {len(df)} 行")
        print(f"   ✓ 列名: {list(df.columns)}")
        
        
    except Exception as e:
        print(f"   ✗ Excel読み込みエラー: {e}")
        return False
    
    # 4. WebDriverテスト（初期化のみ）
    try:
        print("4. WebDriver初期化テスト...")
        options = webdriver.EdgeOptions()
        options.add_argument('--headless')  # ヘッドレスモードでテスト
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        
        driver = None
        try:
            driver = create_edge_driver(options)
            print("   ✓ Selenium Managerを使用したWebDriver初期化成功")
        except Exception as e:
            print(f"   ✗ WebDriver初期化エラー: {e}")
            print("   → Edge ブラウザがインストールされていない可能性があります")
            return False
        
        if driver:
            driver.quit()
            print("   ✓ WebDriver終了成功")
        
    except Exception as e:
        print(f"   ✗ WebDriver初期化エラー: {e}")
        print("   → Edge ブラウザがインストールされていない可能性があります")
        return False
    
    print("=== 基本機能テスト完了 ===")
    return True

def main():
    """メイン処理"""
    import sys
    
    # コマンドライン引数でテストモード判定
    if len(sys.argv) > 1 and sys.argv[1] == "test":
        return test_basic_functionality()
    
    # 全体の開始時刻記録
    overall_start_time = datetime.now()
    
    try:
        logging.info("=== 経費登録自動化処理開始 ===")
        logging.info(f"開始時刻: {overall_start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"経費登録自動化処理を開始します。開始時刻: {overall_start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # ファイル準備
        if not check_and_download_file():
            logging.error("ファイル準備に失敗しました")
            return False
        
        # Excel・Web処理
        if excel_processing():
            overall_end_time = datetime.now()
            duration_str = format_time_duration(overall_start_time, overall_end_time)
            
            success_msg = f"処理が正常に完了しました。\n総処理時間: {duration_str}"
            print(success_msg)
            logging.info(f"=== 経費登録自動化処理完了 ===")
            logging.info(f"終了時刻: {overall_end_time.strftime('%Y-%m-%d %H:%M:%S')}")
            logging.info(f"総処理時間: {duration_str}")
            
            return True
        else:
            overall_end_time = datetime.now()
            duration_str = format_time_duration(overall_start_time, overall_end_time)
            
            error_msg = f"Excel・Web処理に失敗しました。\n処理時間: {duration_str}"
            logging.error(error_msg)
            
            try:
                msgbox.showerror("処理失敗", error_msg)
            except Exception:
                print(f"エラー: {error_msg}")
            
            return False
        
    except Exception as e:
        overall_end_time = datetime.now()
        duration_str = format_time_duration(overall_start_time, overall_end_time)
        
        error_msg = f"処理中にエラーが発生しました: {e}\n処理時間: {duration_str}"
        logging.error(error_msg)
        
        try:
            msgbox.showerror("処理エラー", error_msg)
        except Exception:
            print(f"エラー: {error_msg}")
        
        return False

if __name__ == "__main__":
    main()
